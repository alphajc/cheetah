import os
import shutil
import openpyxl
import logging
import pandas as pd
from xlrd.biffh import XLRDError
from datetime import date
from dateutil.relativedelta import relativedelta


class Point:
    def __init__(self, datadir, config):
        self.points_fpath = os.path.join(datadir, config['points_file'])
        self.config = config
        self.logger = logging.getLogger(__name__)

        self.total_points_list = pd.read_excel(
            self.points_fpath,
            sheet_name='total',
            index_col='悦跑圈ID'
        )
        self.monthly_points_list = pd.read_excel(
            self.points_fpath,
            sheet_name='monthly',
            index_col='悦跑圈ID'
        )

    """加分"""
    def award(self, records_fpath, get_points):
        try:
            records = pd.read_excel(records_fpath, header=10, index_col='用户ID')
        except XLRDError:
            self.logger.info('昨日无跑步记录')
            return

        for record in records.itertuples():
            new_points = get_points(record[-3:])  # 计算新积分
            name = record[1].replace(self.config['name_prefix'], '')
            if new_points == 0:
                continue
            self.logger.info(
                '{}有{}次新跑步记录，累计{}公里，用时{}'.format(
                    name, record[-1], record[-3], record[-2]))
            try:
                self.total_points_list \
                    .loc[record.Index, '累计分'] += new_points
                self.total_points_list \
                    .loc[record.Index, '当前分'] += new_points
            except KeyError:
                """首次统计到的队员"""
                new_points += 2  # 新队员首跑加2分
                self.total_points_list.loc[record.Index] = {
                    '姓名': name,
                    '累计分': new_points,
                    '已消费': 0,
                    '当前分': new_points
                }

            try:
                self.monthly_points_list \
                    .loc[record.Index, '月度积分'] += new_points
            except KeyError:
                total_record = self.total_points_list.loc[record.Index]
                self.monthly_points_list.loc[record.Index] = {
                    '姓名': total_record['姓名'],
                    '月度积分': new_points
                }

    """滚动月榜"""
    def roll_monthly_list(self):
        last_month = date.today() - relativedelta(months=1)
        workbook = openpyxl.load_workbook(self.points_fpath)
        old_sheet = workbook['monthly']
        old_sheet.title = '{}年{}月积分榜'.format(
            last_month.year,
            last_month.month
        )
        new_sheet = workbook.create_sheet('monthly')
        new_sheet['A1'] = '姓名'
        new_sheet['B1'] = '悦跑圈ID'
        new_sheet['C1'] = '月度积分'
        workbook.save(self.points_fpath)
        self.logger.info('滚动月榜')

        """重读'monthly'"""
        self.monthly_points_list = pd.read_excel(
            self.points_fpath,
            sheet_name='monthly',
            index_col='悦跑圈ID'
        )

    """持久化"""
    def persist(self):
        with pd.ExcelWriter(
            self.points_fpath,
            mode='a',
            engine='openpyxl'
        ) as writer:
            """备份"""
            shutil.copyfile(
                writer.path,
                os.path.join(
                    os.path.dirname(writer.path),
                    '{}_{}'.format(
                        date.today().strftime('%Y%m%d'),
                        os.path.basename(writer.path)
                    )
                )
            )
            writer.book.remove(writer.book['total'])
            writer.book.remove(writer.book['monthly'])
            writer.book.create_sheet('total', 0)
            writer.book.create_sheet('monthly', 1)
            writer.sheets = dict((i.title, i) for i in writer.book.worksheets)
            self.total_points_list.sort_values(
                by=['累计分'], ascending=False, inplace=True)
            self.total_points_list.to_excel(
                writer, sheet_name='total')
            self.monthly_points_list.sort_values(
                by=['月度积分'], ascending=False, inplace=True)
            self.monthly_points_list.to_excel(
                writer, sheet_name='monthly')
            writer.sheets['total'].column_dimensions['A'].width = 12
            writer.sheets['monthly'].column_dimensions['A'].width = 12

        self.logger.info('新积分已写入文件')
